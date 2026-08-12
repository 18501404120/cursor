#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按已锁定口径，从业务线下表重建原型 base.js 与初始化 xlsx。"""

from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from copy import deepcopy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DED = ROOT / "原型/02-费用管理/商超计提表格/商超促销扣款&销售折扣&现金折扣&销售费用测算数-2026年(2).xlsx"
REF = ROOT / "原型/02-费用管理/商超计提表格/商超退货扣款计提测算-2026年@0506.xlsx"
BASE_JS = ROOT / "原型/02-费用管理/supermarket-accrual-base.js"
OUT_XLSX = Path(__file__).resolve().parent / "费用管理-客户计提规则退款扣款-初始化模板.xlsx"

PERIODS = ["2026-01", "2026-02", "2026-03", "2026-04"]
FEE_DEDUCTION = ["促销扣款", "销售折扣", "现金折扣", "销售费用"]

CUSTOMERS_24 = [
    "AAFES", "B&H", "Best Buy", "Best Buy 3P", "Costco CA", "Costco US",
    "D&H CA", "D&H US", "HOME DEPOT CANADA", "Kohl's", "LOEWS 3P-US", "Lowe's",
    "Meijer", "Menards", "Micro Center", "Nexcom", "Sam's Club", "Staples",
    "Synnex-CA", "Synnex-US", "Target", "The Home Depot", "Walmart 3P", "Walmart-DSV",
]

# 退款币种 / 扣款币种
CURRENCY = {
    "Costco CA": {"refund": "CNY", "deduction": "USD"},
    "D&H CA": {"refund": "CNY", "deduction": "CAD"},
    "HOME DEPOT CANADA": {"refund": "CNY", "deduction": "CAD"},
    "Synnex-CA": {"refund": "CNY", "deduction": "CAD"},
}
DEFAULT_CUR = {"refund": "USD", "deduction": "USD"}

NAME_ALIAS = {
    "Costco": "Costco US",
    "COSTCO CA": "Costco CA",
    "Walmart": "Walmart-DSV",
    "B&H PHOTO - VIDEO, INC.": "B&H",
    "SYNNEX CA": "Synnex-CA",
    "Synnex": "Synnex-US",
}


def norm_name(name):
    if not name or not isinstance(name, str):
        return None
    name = name.strip()
    if name in ("合计", "参数", "客户补发") or "汇总" in name:
        return None
    return NAME_ALIAS.get(name, name)


def ym(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m")
    s = str(v).strip()
    if "汇总" in s:
        return None
    m = re.match(r"(\d{4}-\d{2})", s)
    return m.group(1) if m else None


def r2(x):
    if x is None:
        return 0.0
    try:
        return round(float(x), 2)
    except Exception:
        return 0.0


def r8(x):
    if x is None:
        return 0.0
    try:
        return round(float(x), 8)
    except Exception:
        return 0.0


def num_or_none(x):
    if x is None or x == "" or isinstance(x, str):
        return None
    try:
        v = float(x)
        return v if math.isfinite(v) else None
    except Exception:
        return None


def cur(customer, domain):
    return CURRENCY.get(customer, DEFAULT_CUR)[domain]


def load_existing_base():
    text = BASE_JS.read_text(encoding="utf-8")
    m = re.search(r"window\.SupermarketAccrualBaseData\s*=\s*(\{.*\})\s*;?\s*$", text, re.S)
    return json.loads(m.group(1))


def parse_ratio_sheet(wb):
    ws = wb["扣款比例"]
    hdr = [c.value for c in ws[1]]
    cust_idx = {hdr[i]: i for i in range(5, len(hdr)) if hdr[i]}
    rows = list(ws.iter_rows(min_row=1, max_row=6, values_only=True))
    out = {}
    for c, i in cust_idx.items():
        c = norm_name(c) or c
        promo_h1 = num_or_none(rows[1][i])
        promo_h2 = num_or_none(rows[2][i])
        disc = rows[3][i]
        cash = rows[4][i]
        ads = rows[5][i] if len(rows[5]) > i else None
        disc_n = num_or_none(disc)
        cash_n = num_or_none(cash)
        ads_n = num_or_none(ads)
        out[c] = {
            "promo_h1": promo_h1 or 0.0,
            "promo_h2": promo_h2 or 0.0,
            "disc": disc_n,
            "disc_raw": disc,
            "cash": cash_n if cash_n is not None else 0.0,
            "ads": ads_n,
            "ads_raw": ads,
        }
    return out


def parse_accrual_amount(wb):
    """计提扣款金额右块：期间/客户/收入/四费计提。"""
    ws = wb["计提扣款金额"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        p, c = ym(r[4]), norm_name(r[5])
        if not p or not c:
            continue
        rows.append({
            "period": p,
            "customer": c,
            "income": r2(r[6]),
            "promoAccrual": r2(r[7]) if r[7] is not None else 0.0,
            "salesDiscountAccrual": r2(r[8]) if r[8] is not None else 0.0,
            "cashDiscountAccrual": r2(r[9]) if r[9] is not None else 0.0,
            "salesExpenseAccrual": r2(r[10]) if r[10] is not None else 0.0,
            "note": (r[11] or "") if len(r) > 11 else "",
        })
    return rows


def parse_promo_writeoff(wb):
    ws = wb["促销扣款冲销记录"]
    # 0-based: 16=2025终, then Jan:17 bill,19 accr,20 close; Feb:21 bill,22 accr,24 close; Mar:25/26/28; Apr:29/30/32
    month_map = [
        ("2026-01", 16, 17, 19, 20),  # open from year-end col16
        ("2026-02", 20, 21, 22, 24),
        ("2026-03", 24, 25, 26, 28),
        ("2026-04", 28, 29, 30, 32),
    ]
    out = []
    for r in ws.iter_rows(min_row=9, values_only=True):
        c = norm_name(r[3])
        if not c:
            continue
        for i, (period, open_i, bill_i, accr_i, close_i) in enumerate(month_map):
            opening = r2(r[open_i]) if r[open_i] is not None else 0.0
            actual = r2(r[bill_i]) if bill_i < len(r) and r[bill_i] is not None else 0.0
            accr = r2(r[accr_i]) if accr_i < len(r) and r[accr_i] is not None else 0.0
            closing = r2(r[close_i]) if close_i < len(r) and r[close_i] is not None else r2(opening + accr - actual)
            out.append({
                "feeType": "促销扣款",
                "customer": c,
                "period": period,
                "openingBalance": opening,
                "actualAmount": actual,
                "excelAccrual": accr,
                "closingBalance": closing,
            })
    return out


def parse_wide_writeoff(ws, fee_type, start_row=3):
    """销折/现折/销费：G年终, H账单,J预估,K余额 | L账单,N预估,O余额 | P账单,R预估,S余额 | T账单,V预估,W余额
    现金折扣列顺序略有差异：看 R2。销费用 A客户 B年终, C账单,E预估,F余额 ...
    """
    # detect by fee
    if fee_type == "销售费用":
        # A=cust, B=yearend, then blocks of 账单/判断/预估/余额 starting C
        # C bill, E accr, F close; G bill, I accr, J close; K bill, M accr, N close; O bill, Q accr, R close
        blocks = [
            ("2026-01", 1, 2, 4, 5),
            ("2026-02", 5, 6, 8, 9),
            ("2026-03", 9, 10, 12, 13),
            ("2026-04", 13, 14, 16, 17),
        ]
        name_i = 0
    elif fee_type == "现金折扣":
        # from earlier: G yearend, H bill, J accr, K close; L bill, M accr(!), O close; P bill, Q accr, S close; T bill...
        # R2: 账单,判断,预估,余额 then 账单,预估,判断,余额 — Feb order differs
        blocks = [
            ("2026-01", 6, 7, 9, 10),
            ("2026-02", 10, 11, 12, 14),  # L bill, M accr, O close (N=判断)
            ("2026-03", 14, 15, 16, 18),
            ("2026-04", 18, 19, 20, 22),
        ]
        name_i = 0
    else:  # 销售折扣
        blocks = [
            ("2026-01", 6, 7, 9, 10),
            ("2026-02", 10, 11, 13, 14),
            ("2026-03", 14, 15, 17, 18),
            ("2026-04", 18, 19, 21, 22),
        ]
        name_i = 0

    out = []
    for r in ws.iter_rows(min_row=start_row, values_only=True):
        c = norm_name(r[name_i])
        if not c:
            continue
        for period, open_i, bill_i, accr_i, close_i in blocks:
            opening = r2(r[open_i]) if open_i < len(r) and r[open_i] is not None else 0.0
            actual = r2(r[bill_i]) if bill_i < len(r) and r[bill_i] is not None else 0.0
            accr = r2(r[accr_i]) if accr_i < len(r) and r[accr_i] is not None else 0.0
            if close_i < len(r) and r[close_i] is not None:
                closing = r2(r[close_i])
            else:
                closing = r2(opening + accr - actual)
            out.append({
                "feeType": fee_type,
                "customer": c,
                "period": period,
                "openingBalance": opening,
                "actualAmount": actual,
                "excelAccrual": accr,
                "closingBalance": closing,
            })
    return out


def parse_refund_writeoff(wb):
    ws = wb["销售退款计提冲销记录"]
    # B yearend, C Jan bill, E Jan accr?, F Jan close, G Feb bill, H Feb accr, J Feb close, ...
    # From earlier: E is formula for Jan accrual = C+(F-B)
    blocks = [
        ("2026-01", 1, 2, 4, 5),
        ("2026-02", 5, 6, 7, 9),
        ("2026-03", 9, 10, 11, 13),
        ("2026-04", 13, 14, 15, 17),
    ]
    out = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        c = norm_name(r[0])
        if not c:
            continue
        for period, open_i, bill_i, accr_i, close_i in blocks:
            opening = r2(r[open_i]) if open_i < len(r) and r[open_i] is not None else 0.0
            actual = r2(r[bill_i]) if bill_i < len(r) and r[bill_i] is not None else 0.0
            if accr_i < len(r) and r[accr_i] is not None:
                accr = r2(r[accr_i])
            else:
                # derive from closing if present
                closing_tmp = r2(r[close_i]) if close_i < len(r) and r[close_i] is not None else None
                accr = r2(closing_tmp - opening + actual) if closing_tmp is not None else 0.0
            closing = r2(r[close_i]) if close_i < len(r) and r[close_i] is not None else r2(opening + accr - actual)
            out.append({
                "customer": c,
                "period": period,
                "openingBalance": opening,
                "actualAmount": actual,
                "excelAccrual": accr,
                "closingBalance": closing,
            })
    return out


def effective_disc_ratio(accrual_rows, customer):
    for p in PERIODS:
        for r in accrual_rows:
            if r["customer"] == customer and r["period"] == p and r["income"] > 0:
                return r8(r["salesDiscountAccrual"] / r["income"])
    return 0.0


def build_fixed_rules(ratios, accrual_rows, old_base):
    """规则口径以客户计提规则页面截图为准（2026-08-10）。"""
    old_map = {r["id"]: r for r in old_base.get("fixedRules", [])}

    # Target / Walmart-DSV 销折部门明细（对应部门管理截图）
    dept_sales_discount = {
        "Target|销售折扣": [
            {"id": "207", "code": "207", "name": "D207", "ratio": 0.046},
            {"id": "008", "code": "008", "name": "D8", "ratio": 0.045},
        ],
        "Walmart-DSV|销售折扣": [
            {"id": "00072", "code": "00072", "name": "D72", "ratio": 0.0414},
            {"id": "00016", "code": "00016", "name": "D16", "ratio": 0.0209},
            {"id": "00011", "code": "00011", "name": "D11", "ratio": 0.0359},
        ],
    }
    for key, fallback in list(dept_sales_discount.items()):
        kept = old_map.get(key, {}).get("deptItems")
        if kept:
            dept_sales_discount[key] = kept

    annual_expense = {
        "Best Buy": 674900.04,
        "Lowe's": 654999.96,
        "The Home Depot": 755000.04,
        "Walmart-DSV": 1269999.96,
    }

    rules = []
    for c in CUSTOMERS_24:
        rt = ratios.get(c, {"promo_h1": 0, "promo_h2": 0, "disc": 0, "cash": 0})
        # sample base amounts from 2026-01 accrual if any
        sample = next((x for x in accrual_rows if x["customer"] == c and x["period"] == "2026-01"), None)
        if not sample:
            sample = next((x for x in accrual_rows if x["customer"] == c and x["period"] in PERIODS), None)

        def sample_amt(field):
            return r2(sample[field]) if sample else 0.0

        # 促销（Target 页面口径为固定 8.49%，保留 H2）
        promo_ratio = 0.0849 if c == "Target" else r8(rt["promo_h1"])
        promo = {
            "id": f"{c}|促销扣款",
            "customer": c,
            "feeType": "促销扣款",
            "method": "fixed_ratio",
            "ratio": promo_ratio,
            "ratioH2": r8(rt["promo_h2"]),
            "baseAmount": sample_amt("promoAccrual"),
            "sourcePeriod": "2026-01" if c != "HOME DEPOT CANADA" else "2026-02",
            "note": (
                "页面口径：固定比例 8.49%；H2 点位见 ratioH2"
                if c == "Target"
                else "促销比例按期间切换：1-6月用H1，7-12月用H2"
            ),
        }
        rules.append(promo)

        # 销售折扣
        if c == "Best Buy":
            disc = {
                "id": f"{c}|销售折扣",
                "customer": c,
                "feeType": "销售折扣",
                "method": "fixed_ratio",
                "ratio": 0.105,
                "baseAmount": sample_amt("salesDiscountAccrual"),
                "sourcePeriod": "2026-01",
                "note": "页面口径：固定比例 10.50%（不再用部门 7%+3.5%）",
            }
        elif c in ("Target", "Walmart-DSV"):
            disc = {
                "id": f"{c}|销售折扣",
                "customer": c,
                "feeType": "销售折扣",
                "method": "dept_fixed_ratio",
                "ratio": None,
                "baseAmount": sample_amt("salesDiscountAccrual"),
                "sourcePeriod": "2026-01",
                "note": (
                    "部门固定比例：D207 4.6% | D8 4.5%"
                    if c == "Target"
                    else "部门固定比例：D72 4.14% | D16 2.09% | D11 3.59%"
                ),
                "deptItems": dept_sales_discount[f"{c}|销售折扣"],
            }
        else:
            disc_ratio = rt["disc"]
            if disc_ratio is None:
                disc_ratio = effective_disc_ratio(accrual_rows, c)
            disc = {
                "id": f"{c}|销售折扣",
                "customer": c,
                "feeType": "销售折扣",
                "method": "fixed_ratio",
                "ratio": r8(disc_ratio or 0),
                "baseAmount": sample_amt("salesDiscountAccrual"),
                "sourcePeriod": "2026-01" if c != "HOME DEPOT CANADA" else "2026-02",
                "note": "",
            }
        rules.append(disc)

        # 现金折扣
        rules.append({
            "id": f"{c}|现金折扣",
            "customer": c,
            "feeType": "现金折扣",
            "method": "fixed_ratio",
            "ratio": r8(rt.get("cash") or 0),
            "baseAmount": sample_amt("cashDiscountAccrual"),
            "sourcePeriod": "2026-01" if c != "HOME DEPOT CANADA" else "2026-02",
            "note": "",
        })

        # 销售费用
        if c == "Sam's Club":
            se = {
                "id": f"{c}|销售费用",
                "customer": c,
                "feeType": "销售费用",
                "method": "fixed_ratio",
                "ratio": 0.01,
                "baseAmount": sample_amt("salesExpenseAccrual"),
                "sourcePeriod": "2026-01",
                "note": "收入×1%",
            }
        elif c in annual_expense:
            se = {
                "id": f"{c}|销售费用",
                "customer": c,
                "feeType": "销售费用",
                "method": "annual_avg",
                "ratio": 0,
                "baseAmount": annual_expense[c],
                "sourcePeriod": "2026-01",
                "note": f"年框 {annual_expense[c]}/12",
            }
        else:
            # Costco 等强制 0；其余无源也 0
            se = {
                "id": f"{c}|销售费用",
                "customer": c,
                "feeType": "销售费用",
                "method": "budget_or_fixed",
                "ratio": 0,
                "baseAmount": 0,
                "sourcePeriod": "2026-01",
                "note": "",
            }
        rules.append(se)
    return rules


def ensure_full_deduction_ledgers(parsed_rows):
    idx = {(r["customer"], r["feeType"], r["period"]): r for r in parsed_rows}
    out = []
    for c in CUSTOMERS_24:
        for fee in FEE_DEDUCTION:
            prev_close = None
            for i, p in enumerate(PERIODS):
                key = (c, fee, p)
                if key in idx:
                    row = dict(idx[key])
                    if prev_close is not None:
                        # 强制期初衔接
                        row["openingBalance"] = r2(prev_close)
                        row["closingBalance"] = r2(row["openingBalance"] + row["excelAccrual"] - row["actualAmount"])
                    prev_close = row["closingBalance"]
                else:
                    opening = r2(prev_close) if prev_close is not None else 0.0
                    row = {
                        "feeType": fee,
                        "customer": c,
                        "period": p,
                        "openingBalance": opening,
                        "actualAmount": 0.0,
                        "excelAccrual": 0.0,
                        "closingBalance": opening,
                    }
                    prev_close = opening
                row["id"] = f"{fee}|{c}|{p}"
                out.append(row)
    return out


def ensure_full_refund_ledgers(parsed_rows):
    idx = {(r["customer"], r["period"]): r for r in parsed_rows}
    out = []
    for c in CUSTOMERS_24:
        prev_close = None
        for p in PERIODS:
            if (c, p) in idx:
                row = dict(idx[(c, p)])
                if prev_close is not None:
                    row["openingBalance"] = r2(prev_close)
                    row["closingBalance"] = r2(row["openingBalance"] + row["excelAccrual"] - row["actualAmount"])
                prev_close = row["closingBalance"]
            else:
                opening = r2(prev_close) if prev_close is not None else 0.0
                row = {
                    "customer": c,
                    "period": p,
                    "openingBalance": opening,
                    "actualAmount": 0.0,
                    "excelAccrual": 0.0,
                    "closingBalance": opening,
                }
                prev_close = opening
            row["id"] = f"{c}|{p}"
            out.append(row)
    return out


def build_refund_rules(ledgers, old_base):
    old = {(r["customer"], r["period"]): r for r in old_base.get("refundRules", [])}
    out = []
    for row in ledgers:
        c, p = row["customer"], row["period"]
        prev = old.get((c, p), {})
        ratio = prev.get("ratio")
        window = prev.get("windowMonths", 3)
        sales_basis = prev.get("salesBasis")
        target = row["closingBalance"]
        if ratio is None:
            ratio = 0.0
        if not sales_basis:
            sales_basis = r2(target / ratio) if ratio else 0.0
        out.append({
            "id": f"{c}|{p}",
            "customer": c,
            "period": p,
            "feeType": "销售退款",
            "method": "rolling_refund",
            "ratio": r8(ratio),
            "windowMonths": window,
            "salesBasis": r2(sales_basis),
            "targetClosing": r2(target),
        })
    return out


def build_accrual_rows(parsed, old_base):
    """合并计提扣款金额 + 补全客户期间；保留历史月来自 old refund/accrual。"""
    by = {(r["period"], r["customer"]): r for r in parsed}
    # keep non-PERIODS historical from old if any in accrualRows
    out = []
    seen = set()
    for r in old_base.get("accrualRows", []):
        c = norm_name(r.get("customer")) or r.get("customer")
        p = r.get("period")
        if not c or not p:
            continue
        if p in PERIODS or p == "2025-12":
            # prefer fresh parsed when available
            if (p, c) in by:
                nr = dict(by[(p, c)])
                out.append(nr)
                seen.add((p, c))
                continue
        row = dict(r)
        row["customer"] = c
        out.append(row)
        seen.add((p, c))
    for r in parsed:
        key = (r["period"], r["customer"])
        if key not in seen:
            out.append(r)
            seen.add(key)
    # ensure all customers have PERIODS (+ keep 2025-12 if exists)
    for c in CUSTOMERS_24:
        for p in ["2025-12"] + PERIODS:
            if (p, c) not in seen:
                out.append({
                    "period": p,
                    "customer": c,
                    "income": 0.0,
                    "promoAccrual": 0.0,
                    "salesDiscountAccrual": 0.0,
                    "cashDiscountAccrual": 0.0,
                    "salesExpenseAccrual": 0.0,
                    "note": "",
                })
                seen.add((p, c))
    out.sort(key=lambda x: (x["period"], x["customer"]))
    return out


def extend_refund_history(old_hist):
    hist = {}
    for k, v in old_hist.items():
        nk = norm_name(k) or k
        hist[nk] = deepcopy(v)
    for c in CUSTOMERS_24:
        if c not in hist:
            # minimal zero history for rolling window
            hist[c] = [{"period": f"2025-{m:02d}", "sales": 0.0, "refund": 0.0} for m in range(1, 13)]
            for p in PERIODS:
                hist[c].append({"period": p, "sales": 0.0, "refund": 0.0})
    return hist


def write_base_js(data):
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    BASE_JS.write_text(f"window.SupermarketAccrualBaseData = {payload};\n", encoding="utf-8")


def style_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="D9E2F3")
    font = Font(bold=True)
    for i in range(1, cols + 1):
        cell = ws.cell(row=row, column=i)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def write_xlsx(base):
    wb = Workbook()
    # 00
    ws = wb.active
    ws.title = "00-填写说明"
    ws["A1"] = "费用管理 · 客户计提规则 / 退款管理 / 扣款管理 · 初始化数据模板"
    ws["A1"].font = Font(bold=True, size=14)
    notes = [
        ("模板定位", "上线初始化导入模板。数值按业务线下测算表重建（对齐口径 2026-08）。"),
        ("客户范围", "全量 24 客户；无源数据补 0 规则/0 台账。"),
        ("币种", "退款/负债：加拿大客户 CNY；扣款：Costco CA=USD，D&H CA/HD CA/Synnex-CA=CAD，其余 USD。同一客户可按费用域拆币种。"),
        ("促销 H1/H2", "促销扣款按期间切换：1-6 月 H1，7-12 月 H2（规则含 ratio 与 ratioH2）。"),
        ("销售费用", "Sam's=收入×1%；Best Buy/Lowe's/THD/Walmart-DSV=年框÷12；Costco 销售费用=0；其余 0。"),
        ("部门比例 D", "Target|销售折扣、Walmart-DSV|销售折扣；部门主数据见 D0（Target/Walmart-DSV）。"),
        ("匹配键", "退款：客户+期间；扣款：客户+费用项+期间；规则：客户+费用项；部门：客户+费用项+部门编码；收入：客户+期间。"),
        ("参考列", "带【参考】前缀为系统可计算/校验参考，不建议覆盖导入。"),
    ]
    ws["A3"] = "项"
    ws["B3"] = "说明"
    style_header(ws, 3, 2)
    for i, (a, b) in enumerate(notes, 4):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 90

    # 99
    ws = wb.create_sheet("99-枚举与映射")
    ws["A1"] = "枚举与字段映射（上线研发对照）"
    ws["A3"] = "计提方式（C）"
    ws.append([])
    for row in [
        ["页面/导入值", "原型码", "说明"],
        ["固定比例", "fixed_ratio", "计提用收入 × 比例（促销可按期间取 H1/H2）"],
        ["部门固定比例", "dept_fixed_ratio", "各部门订单金额×部门比例；明细见 D"],
        ["月固定金额", "budget_or_fixed / monthly_fixed", "基数/月固定金额"],
        ["年总额月均", "annual_avg", "年框÷12"],
    ]:
        ws.append(row)
    ws.append([])
    ws.append(["币种域"])
    ws.append(["客户", "退款币种", "扣款币种"])
    for c in CUSTOMERS_24:
        cc = CURRENCY.get(c, DEFAULT_CUR)
        ws.append([c, cc["refund"], cc["deduction"]])

    # A
    ws = wb.create_sheet("A-客户清单")
    ws["A1"] = "A · 客户清单（导入主数据）"
    ws["A2"] = "扣款币种见「扣款币种」列；退款币种见「退款币种」列。"
    ws["A4"] = "客户"
    headers = ["客户", "退款币种", "扣款币种", "是否启用", "数据性质", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(headers))
    for i, c in enumerate(CUSTOMERS_24, 5):
        cc = CURRENCY.get(c, DEFAULT_CUR)
        note = ""
        if c == "Costco CA":
            note = "退款CNY / 扣款USD"
        elif c in CURRENCY:
            note = f"退款{cc['refund']} / 扣款{cc['deduction']}"
        ws.cell(i, 1, c)
        ws.cell(i, 2, cc["refund"])
        ws.cell(i, 3, cc["deduction"])
        ws.cell(i, 4, "是")
        ws.cell(i, 5, "示例")
        ws.cell(i, 6, note)

    # B
    ws = wb.create_sheet("B-关账期间")
    ws["A1"] = "B · 关账期间（全局，期间另定）"
    headers = ["关账期间", "是否关账", "数据性质", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, 4)
    for i, p in enumerate(base.get("closedPeriods", ["2026-01", "2026-02"]), 5):
        ws.cell(i, 1, p)
        ws.cell(i, 2, "是")
        ws.cell(i, 3, "示例")
        ws.cell(i, 4, "原型示例，上线另定")
    ws.cell(7, 1, "2026-03")
    ws.cell(7, 2, "否（最早未关账，参考）")
    ws.cell(7, 3, "示例")
    ws.cell(7, 4, "【参考】最早未关账月")

    # C
    ws = wb.create_sheet("C-客户计提规则")
    ws["A1"] = "C · 客户计提规则（全量含 0）"
    ws["A2"] = "匹配键：客户+费用项。促销含 H1(比例) / H2(比例H2)。"
    headers = ["客户", "费用项", "扣款币种", "计提方式", "比例", "比例H2", "基数/月固定金额或年总额", "生效期间", "数据性质", "备注",
               "【参考】原型规则ID", "【参考】原型method码"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(headers))
    method_label = {
        "fixed_ratio": "固定比例",
        "dept_fixed_ratio": "部门固定比例",
        "budget_or_fixed": "月固定金额",
        "monthly_fixed": "月固定金额",
        "annual_avg": "年总额月均",
    }
    r = 5
    for rule in base["fixedRules"]:
        ws.cell(r, 1, rule["customer"])
        ws.cell(r, 2, rule["feeType"])
        ws.cell(r, 3, cur(rule["customer"], "deduction"))
        ws.cell(r, 4, method_label.get(rule["method"], rule["method"]))
        ws.cell(r, 5, rule.get("ratio"))
        ws.cell(r, 6, rule.get("ratioH2"))
        ws.cell(r, 7, rule.get("baseAmount"))
        ws.cell(r, 8, rule.get("sourcePeriod"))
        ws.cell(r, 9, "示例")
        ws.cell(r, 10, rule.get("note") or None)
        ws.cell(r, 11, rule["id"])
        ws.cell(r, 12, rule["method"])
        r += 1

    # D
    ws = wb.create_sheet("D-部门固定比例明细")
    ws["A1"] = "D · 部门固定比例明细（仅已能量化拆分的规则）"
    headers = ["客户", "费用项", "扣款币种", "部门编码", "部门名称", "部门比例", "数据性质", "备注", "【参考】原型规则ID"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(headers))
    r = 5
    for rule in base["fixedRules"]:
        if rule.get("method") != "dept_fixed_ratio":
            continue
        for d in rule.get("deptItems") or []:
            ws.cell(r, 1, rule["customer"])
            ws.cell(r, 2, rule["feeType"])
            ws.cell(r, 3, cur(rule["customer"], "deduction"))
            ws.cell(r, 4, d.get("code") or d.get("id"))
            ws.cell(r, 5, d.get("name"))
            ws.cell(r, 6, d.get("ratio"))
            ws.cell(r, 7, "示例")
            ws.cell(r, 9, rule["id"])
            r += 1

    # E
    ws = wb.create_sheet("E-退款台账基线")
    ws["A1"] = "E · 退款台账基线"
    headers = ["客户", "期间", "退款币种", "期初计提退款余额", "实际退款金额", "数据性质", "备注",
               "【参考】期末计提退款余额", "【参考】原型excelAccrual当月计提", "【参考】原型行ID"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(headers))
    r = 5
    for row in base["refundLedgers"]:
        ws.cell(r, 1, row["customer"])
        ws.cell(r, 2, row["period"])
        ws.cell(r, 3, cur(row["customer"], "refund"))
        ws.cell(r, 4, row["openingBalance"])
        ws.cell(r, 5, row["actualAmount"])
        ws.cell(r, 6, "示例")
        ws.cell(r, 8, row["closingBalance"])
        ws.cell(r, 9, row["excelAccrual"])
        ws.cell(r, 10, row["id"])
        r += 1

    # F
    ws = wb.create_sheet("F-退款滚动规则窗口")
    ws["A1"] = "F · 退款滚动规则窗口"
    headers = ["客户", "期间", "退款币种", "费用项", "计提方式", "目标窗口月数", "数据性质", "备注",
               "【参考】滚动退款率", "【参考】销售基数salesBasis", "【参考】目标期末余额targetClosing", "【参考】原型行ID"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(headers))
    r = 5
    for row in base["refundRules"]:
        ws.cell(r, 1, row["customer"])
        ws.cell(r, 2, row["period"])
        ws.cell(r, 3, cur(row["customer"], "refund"))
        ws.cell(r, 4, "销售退款")
        ws.cell(r, 5, "滚动退款")
        ws.cell(r, 6, row.get("windowMonths", 3))
        ws.cell(r, 7, "示例")
        ws.cell(r, 9, row.get("ratio"))
        ws.cell(r, 10, row.get("salesBasis"))
        ws.cell(r, 11, row.get("targetClosing"))
        ws.cell(r, 12, row["id"])
        r += 1

    # G
    ws = wb.create_sheet("G-扣款台账基线")
    ws["A1"] = "G · 扣款台账基线"
    headers = ["客户", "费用项", "期间", "扣款币种", "期初余额", "实际扣款金额", "数据性质", "备注",
               "【参考】计提扣款", "【参考】期末余额", "【参考】差异(实际-计提)", "【参考】状态", "【参考】原型行ID"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(headers))
    r = 5
    for row in base["deductionLedgers"]:
        diff = r2(row["actualAmount"] - row["excelAccrual"])
        if abs(diff) < 0.02:
            status = "匹配"
        elif row["actualAmount"] > row["excelAccrual"]:
            status = "实际>计提"
        else:
            status = "实际<计提"
        ws.cell(r, 1, row["customer"])
        ws.cell(r, 2, row["feeType"])
        ws.cell(r, 3, row["period"])
        ws.cell(r, 4, cur(row["customer"], "deduction"))
        ws.cell(r, 5, row["openingBalance"])
        ws.cell(r, 6, row["actualAmount"])
        ws.cell(r, 7, "示例")
        ws.cell(r, 9, row["excelAccrual"])
        ws.cell(r, 10, row["closingBalance"])
        ws.cell(r, 11, diff)
        ws.cell(r, 12, status)
        ws.cell(r, 13, row["id"])
        r += 1

    # H
    ws = wb.create_sheet("H-计提用收入")
    ws["A1"] = "H · 计提用收入"
    ws["A2"] = "扣款域收入用扣款币种；退款历史销售额/退款见各客户 refundHistory（退款币种）。"
    headers = ["客户", "期间", "扣款币种", "计提用收入", "数据性质", "备注",
               "【参考】历史销售收入", "【参考】历史实际退款", "【参考】原型促销计提", "【参考】原型销折计提",
               "【参考】原型现折计提", "【参考】原型销费计提", "【参考】数据来源"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(headers))
    hist = base.get("refundHistory", {})
    accr_map = {(r["customer"], r["period"]): r for r in base["accrualRows"]}
    # union periods from hist + accrual
    pairs = set(accr_map.keys())
    for c, rows in hist.items():
        for hr in rows:
            pairs.add((c, hr["period"]))
    r = 5
    for c, p in sorted(pairs, key=lambda x: (x[0], x[1])):
        ar = accr_map.get((c, p), {})
        hr = next((x for x in hist.get(c, []) if x["period"] == p), {})
        income = ar.get("income")
        if income is None:
            income = hr.get("sales", 0)
        ws.cell(r, 1, c)
        ws.cell(r, 2, p)
        ws.cell(r, 3, cur(c, "deduction"))
        ws.cell(r, 4, r2(income))
        ws.cell(r, 5, "示例")
        ws.cell(r, 6, ar.get("note") or None)
        ws.cell(r, 7, r2(hr.get("sales")) if hr else r2(income))
        ws.cell(r, 8, r2(hr.get("refund")) if hr else None)
        if ar:
            ws.cell(r, 9, ar.get("promoAccrual"))
            ws.cell(r, 10, ar.get("salesDiscountAccrual"))
            ws.cell(r, 11, ar.get("cashDiscountAccrual"))
            ws.cell(r, 12, ar.get("salesExpenseAccrual"))
            ws.cell(r, 13, "accrualRows+refundHistory" if hr else "accrualRows")
        else:
            ws.cell(r, 13, "refundHistory")
        r += 1

    wb.save(OUT_XLSX)


def main():
    print("loading existing base...")
    old = load_existing_base()
    print("loading deduction workbook...")
    wb_ded = openpyxl.load_workbook(DED, data_only=True)
    ratios = parse_ratio_sheet(wb_ded)
    accrual_parsed = parse_accrual_amount(wb_ded)
    promo_ledgers = parse_promo_writeoff(wb_ded)
    disc_ledgers = parse_wide_writeoff(wb_ded["销售折扣冲销记录"], "销售折扣")
    cash_ledgers = parse_wide_writeoff(wb_ded["现金折扣冲销记录"], "现金折扣")
    fee_ledgers = parse_wide_writeoff(wb_ded["销售费用计提冲销记录"], "销售费用")
    wb_ded.close()

    print("loading refund workbook...")
    wb_ref = openpyxl.load_workbook(REF, data_only=True)
    refund_ledgers_parsed = parse_refund_writeoff(wb_ref)
    wb_ref.close()

    deduction_all = ensure_full_deduction_ledgers(promo_ledgers + disc_ledgers + cash_ledgers + fee_ledgers)
    refund_all = ensure_full_refund_ledgers(refund_ledgers_parsed)
    fixed_rules = build_fixed_rules(ratios, accrual_parsed, old)
    accrual_rows = build_accrual_rows(accrual_parsed, old)
    refund_rules = build_refund_rules(refund_all, old)
    refund_history = extend_refund_history(old.get("refundHistory", {}))

    # sync refund history actuals from refund ledgers for PERIODS
    for row in refund_all:
        c, p = row["customer"], row["period"]
        lst = refund_history.setdefault(c, [])
        found = False
        for hr in lst:
            if hr["period"] == p:
                # keep sales if present; set refund as negative of actual when actual>0
                if row["actualAmount"]:
                    hr["refund"] = -abs(r2(row["actualAmount"]))
                found = True
                break
        if not found:
            sales = next((a["income"] for a in accrual_rows if a["customer"] == c and a["period"] == p), 0.0)
            lst.append({
                "period": p,
                "sales": r2(sales),
                "refund": -abs(r2(row["actualAmount"])) if row["actualAmount"] else 0.0,
            })
        lst.sort(key=lambda x: x["period"])

    customer_currencies = {c: CURRENCY.get(c, DEFAULT_CUR) for c in CUSTOMERS_24}

    customer_dept_master = {
        "Target": [
            {"id": "207", "code": "207", "name": "D207"},
            {"id": "008", "code": "008", "name": "D8"},
        ],
        "Walmart-DSV": [
            {"id": "00072", "code": "00072", "name": "D72"},
            {"id": "00016", "code": "00016", "name": "D16"},
            {"id": "00011", "code": "00011", "name": "D11"},
        ],
    }

    new_base = {
        "periods": PERIODS,
        "ledgerPeriods": old.get("ledgerPeriods", {
            "2026-01": "46023", "2026-02": "46054", "2026-03": "46082", "2026-04": "46113"
        }),
        "customers": CUSTOMERS_24,
        "customerCurrencies": customer_currencies,
        "customerDeptMaster": customer_dept_master,
        "feeTypes": ["销售退款", "促销扣款", "销售折扣", "现金折扣", "销售费用"],
        "accrualRows": accrual_rows,
        "fixedRules": fixed_rules,
        "refundRules": refund_rules,
        "refundHistory": refund_history,
        "refundLedgers": refund_all,
        "deductionLedgers": deduction_all,
        "closedPeriods": old.get("closedPeriods", ["2026-01", "2026-02"]),
        "demoAnchorPeriod": old.get("demoAnchorPeriod", "2026-04"),
        "erpDeptMaster": old.get("erpDeptMaster", {}),
        "erpOrderDeptAmounts": old.get("erpOrderDeptAmounts", {}),
        "alignmentNotes": {
            "customerScope": 24,
            "ruleSource": "客户计提规则管理页面截图 2026-08-10",
            "promoHalf": "H1 months 01-06, H2 months 07-12",
            "salesExpense": "Sam's fixed_ratio 1%; BestBuy/Lowes/THD/Walmart-DSV annual_avg; Costco=0",
            "bestBuySalesDiscount": "fixed_ratio 10.50%",
            "targetPromo": "fixed_ratio 8.49%",
            "deptRules": ["Target|销售折扣", "Walmart-DSV|销售折扣"],
            "customerDeptMaster": "Target 207/D207+008/D8; Walmart-DSV 00072/D72+00016/D16+00011/D11",
            "currency": "refund CNY for CA; Costco CA deduction USD; other CA deduction CAD; refund UI must use ¥",
        },
    }

    print("writing base.js...")
    write_base_js(new_base)
    print("writing xlsx...")
    write_xlsx(new_base)

    # quick validate
    g_fail = 0
    by = defaultdict(list)
    for r in deduction_all:
        by[(r["customer"], r["feeType"])].append(r)
    for k, rows in by.items():
        rows = sorted(rows, key=lambda x: x["period"])
        for i, r in enumerate(rows):
            calc = r2(r["openingBalance"] + r["excelAccrual"] - r["actualAmount"])
            if abs(calc - r2(r["closingBalance"])) > 0.05:
                g_fail += 1
            if i and abs(r2(rows[i - 1]["closingBalance"]) - r2(r["openingBalance"])) > 0.05:
                g_fail += 1
    print(f"done. customers={len(CUSTOMERS_24)} rules={len(fixed_rules)} G={len(deduction_all)} E={len(refund_all)} G_roll_issues={g_fail}")
    print(f"base -> {BASE_JS}")
    print(f"xlsx -> {OUT_XLSX}")


if __name__ == "__main__":
    main()
